import locale
import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import os
import plotly.express as px
from openpyxl.styles import PatternFill

# Configurações gerais
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DADOS_MONITORING_PATH = os.path.join(BASE_DIR, '.database', 'DATABASE.xlsx')
DADOS_ALMOXARIFADO_PATH = os.path.join(BASE_DIR, '.database', 'Novembro-2024.xlsx')

locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')
locale.atof = lambda x: float(x.replace('.', '').replace(',', '.'))

class Material:
    def __init__(self, nome, cor=None):
        self.nome = nome.strip()
        self.cor = cor

class Estoque:
    def __init__(self, path_almoxarifado):
        self.df_almoxarifado = pd.read_excel(path_almoxarifado, sheet_name=0, header=1)
        self.df_almoxarifado.columns = self.df_almoxarifado.columns.str.strip()
    
    def obter_quantidade(self, nome_material, data_coluna):
        df_filtrado = self.df_almoxarifado.set_index("Produto")
        try:
            return df_filtrado.at[nome_material, data_coluna]
        except KeyError:
            return 0

def formatar_valores(valor):
    """ Formatar valores numéricos no formato 10.000,00 """
    return locale.format_string('%.2f', valor / 1000, grouping=True)  # Dividimos por 1000 para mostrar em milhares

# Função para carregar dados de produção com leitura de cores de células na coluna "DESCRIÇÃO"
def carregar_dados_monitoring():
    wb = load_workbook(DADOS_MONITORING_PATH, data_only=True)
    ws = wb['ProgramaExtrusão']
    dados = pd.read_excel(DADOS_MONITORING_PATH, header=4, sheet_name='ProgramaExtrusão')

    # Processar a coluna de descrição e detectar a cor
    materiais = []
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=2, max_col=2):
        for cell in row:
            # Verificar se a célula contém uma string e não está vazia
            nome_material = cell.value if isinstance(cell.value, str) else str(cell.value) if cell.value else ""
            cor = cell.fill.start_color.index  # Captura o índice da cor
            if cor == 'FFFF00':  # Amarelo
                materiais.append(Material(nome_material, cor='amarelo'))
            elif cor == 'FF0000':  # Vermelho
                materiais.append(Material(nome_material, cor='vermelho'))
            else:
                materiais.append(Material(nome_material))
    
    # Ajustar o comprimento da lista `materiais` para que corresponda ao número de linhas de `dados`
    if len(materiais) > len(dados):
        materiais = materiais[:len(dados)]  # Truncar se `materiais` tiver mais elementos
    elif len(materiais) < len(dados):
        materiais.extend([Material("")] * (len(dados) - len(materiais)))  # Preencher se `materiais` tiver menos elementos

    # Extrair atributos `nome` e `cor` de cada `Material` para colunas separadas
    dados['MaterialNome'] = [m.nome for m in materiais]
    dados['MaterialCor'] = [m.cor for m in materiais]
    
    return dados

# Função para carregar dados do almoxarifado e definir última coluna de dados de estoque
def carregar_dados_almoxarifado():
    df = pd.read_excel(DADOS_ALMOXARIFADO_PATH, sheet_name=0, header=1)
    df.columns = df.columns.str.strip()
    ultima_data = df.columns[-1]  # Última coluna com dados atualizados
    return df, ultima_data

# Comparar demanda de produção com estoque atual
def comparar_demanda_estoque(demanda, estoque):
    # Converter as colunas de demanda e estoque para numérico, substituindo erros por NaN e depois por 0
    demanda = pd.to_numeric(demanda, errors='coerce').fillna(0)
    estoque = pd.to_numeric(estoque, errors='coerce').fillna(0)

    saldo = demanda - estoque
    return pd.DataFrame({
        "Composto": demanda.index,
        "Demanda (kg)": demanda.values,
        "Estoque Atual (kg)": estoque.values,
        "Saldo (kg)": saldo
    })

# Página principal
def pagina3():
    st.header('_Demanda de Polímeros_', divider='gray')
    dados_producao = carregar_dados_monitoring()
    dados_almoxarifado, ultima_data = carregar_dados_almoxarifado()

    if dados_producao is not None and dados_almoxarifado is not None:
        colunas_descritivas = dados_producao.columns[:18]
        colunas_compostos = dados_producao.columns[18:]
        
        # separar dados
        demanda_compostos = dados_producao[colunas_compostos].sum()
        dados_descritivos = dados_producao[colunas_descritivas]
        dados_compostos = dados_producao[colunas_compostos]

        # Exibir informações iniciais
        st.subheader("Informações Gerais dos Produtos")
        st.dataframe(dados_descritivos)

        st.subheader("Demanda geral dos compostos por produto")
        st.dataframe(dados_compostos)

        # Calcular o total de horas de produção
        if "Tot Hrs" in dados_descritivos.columns:
            total_horas = dados_descritivos["Tot Hrs"].sum()
            st.write(f"**Total de Horas de Produção:** {formatar_valores(total_horas)} mil horas")
        
        # elif "Dias" in dados_descritivos.columns:                           (NÃO IMPLEMENTAR NO MOMENTO)
        #     total_dias = dados_descritivos["Dias"].sum()
        #     st.write(f"**Aproximaçao do total de dias planejados:** {formatar_valores(total_dias)}")
        
        # Preparar e filtrar estoque                                    
        estoque = Estoque(DADOS_ALMOXARIFADO_PATH)
        estoque_atual = pd.Series({composto: estoque.obter_quantidade(composto, ultima_data) 
                                   for composto in colunas_compostos}).fillna(0)

        resultado_comparacao = comparar_demanda_estoque(demanda_compostos, estoque_atual)


        compostos_deficit = resultado_comparacao[resultado_comparacao["Saldo (kg)"] < 0]
        if not compostos_deficit.empty:
            st.subheader("Compostos com Necessidade de Reposição")
            st.dataframe(compostos_deficit)

        col1, col2 = st.columns(2)
        with col1:
            st.subheader("Comparativo de Demanda e Estoque")
            st.dataframe(resultado_comparacao)                      # FRACIONAR EM DOIS DATAFRAMES PARA FACILITAR A VISUALIZAÇÃO
        with col2: 
            fig_compostos = px.pie(
            resultado_comparacao,
            names='Composto',
            values='Demanda (kg)',
            title="Distribuição dos Compostos na Produção"
        )
        st.plotly_chart(fig_compostos)
    else:
     st.error("Erro ao carregar os dados da aba 'Programa Extrusão' ou 'Almoxarifado'.")

# Interface do sistema
st.set_page_config(page_title="ds3", page_icon="💭", layout="wide")

imagem_caminho = os.path.join(BASE_DIR, '.uploads', 'Logo.png')
if os.path.exists(imagem_caminho):
    st.sidebar.image(imagem_caminho, use_column_width=True)
else:
    st.sidebar.error(f"Imagem no caminho '{imagem_caminho}' não encontrada.")

if 'pagina_atual' not in st.session_state:
    st.session_state.pagina_atual = 'pagina1'

st.sidebar.markdown("<br><br><br>", unsafe_allow_html=True)
botao_pagina1 = st.sidebar.button('(ICON1)', on_click=lambda: st.session_state.update({'pagina_atual': 'pagina1'}))
st.sidebar.markdown("<br><br><br><br><br><br>", unsafe_allow_html=True)
botao_pagina2 = st.sidebar.button('(ICON2)', on_click=lambda: st.session_state.update({'pagina_atual': 'pagina2'}))
st.sidebar.markdown("<br><br><br><br><br><br>", unsafe_allow_html=True)
botao_pagina3 = st.sidebar.button('(ICON3)', on_click=lambda: st.session_state.update({'pagina_atual': 'pagina3'}))

if st.session_state.pagina_atual == 'pagina1':
    pagina1()
elif st.session_state.pagina_atual == 'pagina2':
    pagina2()
elif st.session_state.pagina_atual == 'pagina3':
    pagina3()